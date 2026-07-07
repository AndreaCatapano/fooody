"""
Genera content-map.xlsx da content-map.csv
- Un foglio per pagina
- Intestazioni colorate (pastello)
- Colonna "Contenuto cliente" evidenziata in giallo/verde
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter

# ── Palette colori pastello per le intestazioni ──────────────────────────────
# Ogni colonna ha il suo colore header
COL_COLORS = {
    'ID':                     'D6E4F0',  # azzurro pastello
    'Sezione':                'D5ECD4',  # verde menta
    'Tipo':                   'FDE8D8',  # pesca
    'Contenuto placeholder attuale': 'F5F0DC',  # avorio
    'Vincoli/Note':           'EDE0F5',  # lavanda
    'Contenuto cliente':      'C8F7C5',  # verde lime brillante
}

# Colore fill per le celle della colonna "Contenuto cliente" (evidenziata)
CLIENT_COL_FILL  = PatternFill('solid', fgColor='EFFBEE')  # verde chiarissimo
CLIENT_HDR_FILL  = PatternFill('solid', fgColor='5CB85C')  # verde deciso

# Intestazione riga sezione (separatore visivo tra sezioni)
SECTION_FILL = PatternFill('solid', fgColor='F0F0F0')

# Font base
FONT_BASE   = Font(name='Arial', size=10)
FONT_BOLD   = Font(name='Arial', size=10, bold=True)
FONT_HEADER = Font(name='Arial', size=10, bold=True, color='1A1A1A')
FONT_ID     = Font(name='Arial', size=10, bold=True, color='555555')
FONT_CLIENT_HDR = Font(name='Arial', size=10, bold=True, color='FFFFFF')

THIN = Side(style='thin', color='DDDDDD')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

WRAP = Alignment(wrap_text=True, vertical='top')
CENTER_V = Alignment(vertical='center')


def page_accent(page):
    accents = {
        'Home':      'E8F4FD',
        'Metodo':    'FEF9E7',
        'Social':    'FDEBD0',
        'Web':       'EAF2FF',
        'Branding':  'F5EEF8',
        'Studio':    'E8F8F5',
        'Contatti':  'FDEDEC',
    }
    return accents.get(page, 'F5F5F5')


def make_sheet(wb, page_name, rows):
    ws = wb.create_sheet(title=page_name)

    headers = ['ID', 'Sezione', 'Tipo',
               'Contenuto placeholder attuale', 'Vincoli/Note',
               'Contenuto cliente']

    col_widths = [8, 22, 18, 50, 36, 42]

    # ── Riga 1: titolo pagina ─────────────────────────────────────────────
    ws.merge_cells('A1:F1')
    title_cell = ws['A1']
    title_cell.value = f'📄  {page_name}'
    title_cell.font = Font(name='Arial', size=13, bold=True, color='1A1A1A')
    title_cell.fill = PatternFill('solid', fgColor=page_accent(page_name).replace('#',''))
    title_cell.alignment = Alignment(vertical='center', indent=1)
    ws.row_dimensions[1].height = 28

    # ── Riga 2: intestazioni colonne ──────────────────────────────────────
    for col_i, (hdr, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=2, column=col_i, value=hdr)
        if hdr == 'Contenuto cliente':
            cell.fill = CLIENT_HDR_FILL
            cell.font = FONT_CLIENT_HDR
        else:
            cell.fill = PatternFill('solid', fgColor=COL_COLORS.get(hdr, 'EEEEEE'))
            cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(col_i)].width = width
    ws.row_dimensions[2].height = 22

    # ── Righe dati ────────────────────────────────────────────────────────
    current_section = None
    data_row = 3

    for row in rows:
        sezione = row.get('Sezione', '')

        # Separatore visivo cambio sezione
        if sezione != current_section:
            current_section = sezione
            ws.merge_cells(
                start_row=data_row, start_column=1,
                end_row=data_row, end_column=6
            )
            sec_cell = ws.cell(row=data_row, column=1,
                               value=f'  {sezione}')
            sec_cell.fill = PatternFill('solid', fgColor='E8E8E8')
            sec_cell.font = Font(name='Arial', size=9, bold=True,
                                 color='555555', italic=True)
            sec_cell.alignment = Alignment(vertical='center', indent=1)
            ws.row_dimensions[data_row].height = 16
            data_row += 1

        values = [
            row.get('ID', ''),
            row.get('Sezione', ''),
            row.get('Tipo', ''),
            row.get('Contenuto placeholder attuale', ''),
            row.get('Vincoli/Note', ''),
            row.get('Contenuto cliente', ''),
        ]

        for col_i, val in enumerate(values, start=1):
            cell = ws.cell(row=data_row, column=col_i, value=val)
            cell.border = BORDER
            cell.alignment = WRAP

            if col_i == 1:  # ID
                cell.font = FONT_ID
                cell.alignment = Alignment(horizontal='center',
                                           vertical='top', wrap_text=False)
            elif col_i == 6:  # Contenuto cliente
                cell.fill = CLIENT_COL_FILL
                cell.font = FONT_BASE
            else:
                cell.font = FONT_BASE

        # Altezza riga adattiva (stima basata sul testo più lungo)
        max_len = max(
            len(str(row.get('Contenuto placeholder attuale', ''))),
            len(str(row.get('Vincoli/Note', '')))
        )
        ws.row_dimensions[data_row].height = max(16, min(80, max_len // 3 + 14))
        data_row += 1

    # ── Freeze intestazioni ───────────────────────────────────────────────
    ws.freeze_panes = 'A3'

    # ── Auto-filter sulla riga 2 ──────────────────────────────────────────
    ws.auto_filter.ref = f'A2:{get_column_letter(len(headers))}2'

    return ws


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    src = '/Users/andreacatapano/Desktop/Lavoro/Fooody/content-map.csv'
    dst = '/Users/andreacatapano/Desktop/Lavoro/Fooody/content-map.xlsx'

    # Leggi CSV
    with open(src, newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        all_rows = list(reader)

    # Raggruppa per pagina
    pages = {}
    for row in all_rows:
        p = row.get('Pagina', 'Altro')
        pages.setdefault(p, []).append(row)

    # Ordine pagine
    page_order = ['Home', 'Metodo', 'Social', 'Web', 'Branding', 'Studio', 'Contatti']
    page_order += [p for p in pages if p not in page_order]

    wb = Workbook()
    wb.remove(wb.active)  # rimuovi foglio vuoto di default

    for page in page_order:
        if page in pages:
            make_sheet(wb, page, pages[page])

    wb.save(dst)
    total = sum(len(v) for v in pages.values())
    print(f'✓ Salvato: {dst}')
    print(f'  {len(pages)} fogli · {total} righe dati')
    for p in page_order:
        if p in pages:
            print(f'  - {p}: {len(pages[p])} elementi')


if __name__ == '__main__':
    main()
